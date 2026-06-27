'''
I learned about openpyxl and how to automate excel with it.
I wanted to just use a .csv file for this but it doesnt work for a more complex data set like the dictionary i am using
I also learnt about Data Frames in pandas, as it is a dictionary written with rows and columns, similar to an excel spreadsheet,
so I used it to upload the data easily from the program to excel.
'''
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from Account import Account
import time
import os
import Bank

file_path = "C:\\Users\\nizat\\Documents\\BankProject\\Book1.xlsx"

bankAccount = None
using = True
loggedIn = False

# Check for the account
def check_account(username, password):
    print("checking for account!")
    df = pd.read_excel(file_path)
    result = df[(df['NAME'] == username) & (df['PASSWORD'] == password)]
    
    # print(result.empty)
    
    if result.empty == True:
        print("Account not found! \n")
        return False
    elif result.empty == False:
        print("Account found! \n")
        return True
    
    

    
# Check for account balance
def check_balance(username, password):
    df = pd.read_excel(file_path)
    result = df[(df['NAME'] == username) & (df['PASSWORD'] == password)]
    resultIndex = result.index.values[0]
    index =  result.index.values[0]  + 1
    
    balance = df.at[index - 1, 'BALANCE'] # learned from https://saturncloud.io/blog/how-to-extract-value-from-a-dataframe-a-comprehensive-guide-for-data-scientists/
    print(balance)
    return balance
# Save account
def save_account_data(account):
    data = account.serialize()
    df = pd.DataFrame(data)
    
    #print(df[['Name']])
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    #print(writer.sheets)
    
    if 'Sheet1' in book.sheetnames:
        workbookdf = pd.read_excel(file_path)
        result = workbookdf[(workbookdf['NAME'] == account.username) & (workbookdf['PASSWORD'] == account.password)]
        if not result.empty:
            '''workbookdf['NAME'] = account.username
            workbookdf['PASSWORD'] = account.password
            workbookdf['BALANCE'] = account.balance'''
            resultIndex = result.index.values[0]
            index =  result.index.values[0]  + 1
            
            df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=index, header=False) # learned from https://www.geeksforgeeks.org/exporting-a-pandas-dataframe-to-an-excel-file/
            writer.close()
        elif result.empty:
            row = book['Sheet1'].max_row if 'Sheet1' in book.sheetnames else 0
            df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=row, header=False) # learned from https://www.geeksforgeeks.org/exporting-a-pandas-dataframe-to-an-excel-file/
            writer.close()
    else:
        print("there was a problem saving your data! :(")
    

# Sign Up/Log in loop
while using:
    print("Welcome to the Roman Bank! \n")
    print("Enter 1, or 2 \n")
    print("1 - Sign Up/Log In")
    print("2 - Exit \n")
    
    choice = int(input("Enter 1, or 2 \n"))
    # Exit
    if choice == 2:
        print("Exiting now...")
        using = False
        break
    # Sign up/Log in
    elif choice == 1:
        username = input("Enter your name \n")
        password = input("Enter your password (make it strong) \n")
        if check_account(username, password): # will return true or false, if true, it logs you in
            print("Logged in!")
            bankAccount = Account(username, password, check_balance(username, password))
            loggedIn = True
            break
        else: # if not, it will create your account and THEN log you in
            confirm_password = input("Enter your password again \n")
                
            if password == confirm_password:
                print("Password is correct! \n")
                print("Signing up... \n")
                bankAccount = Account(username, password, 500)

                save_account_data(bankAccount)
                
                loggedIn = True
                break
                
            else:
                print("Passcode doesn't match")
    time.sleep(0.1)

# Log in loop
while loggedIn:
    print(f"Hello, {bankAccount.username}!")
    print("1 - Check balance")
    print("2 - Deposit")
    print("3 - Withdraw")
    print("4 - Log out")
    choice = int(input("Enter your choice: "))
    
    # Check balance
    if choice == 1:
        print(f"Your balance is: ${bankAccount.balance}")
    # Deposit
    elif choice == 2:
        money = int(input("How much would you like to deposit? (integer or decimal): \n"))
        balance = Bank.deposit(bankAccount.balance,money)
        bankAccount.balance = balance
        print(f"Your balance is: ${bankAccount.balance}")
    
    # Withdraw
    elif choice == 3:
        money = int(input("How much would you like to withdraw? (integer or decimal): \n"))
        if (bankAccount.balance - money) >= 0:
            balance = Bank.withdraw(bankAccount.balance,money)
            bankAccount.balance = balance
            print(f"Your balance is: ${bankAccount.balance}" )
        else:
            print(f"Sorry! You do not have enough money to withdraw ${money}")
    # Log out
    elif choice == 4:
        print("Logging out...")
        save_account_data(bankAccount)            
        loggedIn = False
        break
    time.sleep(0.1)
